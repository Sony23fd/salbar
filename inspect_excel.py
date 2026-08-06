import openpyxl
import json

wb = openpyxl.load_workbook('urtug.xlsx', data_only=True)
wb_formulas = openpyxl.load_workbook('urtug.xlsx', data_only=False)

output = []

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    ws_f = wb_formulas[sheetname]
    
    sheet_data = {
        "sheet": sheetname,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "rows": []
    }
    
    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            formula = ws_f.cell(row=r, column=c).value
            if val is not None or formula is not None:
                row_vals.append({
                    "col": c,
                    "val": str(val) if val is not None else None,
                    "formula": str(formula) if str(formula).startswith('=') else None
                })
        if row_vals:
            sheet_data["rows"].append({
                "row": r,
                "cells": row_vals
            })
            
    output.append(sheet_data)

with open('excel_analysis.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print("Excel analysis saved to excel_analysis.json")
